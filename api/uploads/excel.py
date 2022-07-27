import pandas as pd
from io import BytesIO
from werkzeug.utils import secure_filename
from fastapi.responses import StreamingResponse
from urllib.parse import quote

import openpyxl
import datetime

async def isint(parameter):
    try:
        int(parameter)
        return True
    except ValueError:
        return False

async def isfloat(parameter):
    try:
        float(parameter)
        return True
    except ValueError:
        return False

# テキストファイルをExcelに変換する関数
async def txt_to_xlsx(files):
    col = 0

    # excelファイルの作成
    excel_date = datetime.datetime.today().strftime("%Y%m%d%H%M")
    output_file = excel_date + ".xlsx"

    # openpyxlのセット
    wb = openpyxl.Workbook()
    sheet = wb.active
    xlsx = BytesIO()

    # テキストファイルを読み込む
    for file in files:
        filename = secure_filename(file.filename)
        print(filename)

        # dataframe作成
        col_names = ['c{0:02d}'.format(i) for i in range(10)]
        df = pd.read_csv(file.file, sep='\s+', encoding='SHIFT-JIS', names=col_names)

        # dataframeの必要な部分の抽出
        needed_df = df[["c00", "c01"]]

        # dataframeをexcelに貼り付ける
        for i in range(len(needed_df.columns)):
            for j in range(len(needed_df)):
                column = needed_df.iat[j, i]
                is_int = await(isint(column))
                is_float = await(isfloat(column))
                if is_int:
                    sheet.cell(j+1, col+i+1).value = int(column)
                elif is_float:
                    sheet.cell(j+1, col+i+1).value = float(column)
                else:
                    sheet.cell(j+1, col+i+1).value = column
        
        col = col + 2
    # ファイルの保存
    wb.save(xlsx)
    xlsx.seek(0)
    filename = quote(output_file)
    return StreamingResponse(
        content=xlsx,
        headers={"Content-Disposition": f'attachment; filename="{output_file}"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
