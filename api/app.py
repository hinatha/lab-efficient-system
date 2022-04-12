from natsort import natsorted
import pandas as pd

import os
import openpyxl
import datetime

def isint(parameter):
    try:
        int(parameter)
        return True
    except ValueError:
        return False

def isfloat(parameter):
    try:
        float(parameter)
        return True
    except ValueError:
        return False

# テキストファイルをExcelに変換する関数
def txt2xlsx():
    col = 0
    input_folder = "./input"

    # excelファイルの作成
    excel_date = datetime.datetime.today().strftime("%Y%m%d%H%M")
    output_file = excel_date + ".xlsx"

    # openpyxlのセット
    wb = openpyxl.Workbook()
    sheet = wb.active

    # テキストファイルを読み込む
    for file_name in natsorted(os.listdir(input_folder)):
        print(file_name)
        if not file_name.lower().endswith(".txt"):
            continue
        
        # dataframe作成
        col_names = ['c{0:02d}'.format(i) for i in range(10)]
        file_path = "input/" + file_name
        df = pd.read_csv(file_path, sep='\s+', encoding='SHIFT-JIS', names=col_names)

        # dataframeの必要な部分の抽出
        needed_df = df[["c00", "c01"]]

        # dataframeをexcelに貼り付ける
        for i in range(len(needed_df.columns)):
            for j in range(len(needed_df)):
                column = needed_df.iat[j, i]
                is_int = isint(column)
                is_float = isfloat(column)
                if is_int:
                    sheet.cell(j+1, col+i+1).value = int(column)
                elif is_float:
                    sheet.cell(j+1, col+i+1).value = float(column)
                else:
                    sheet.cell(j+1, col+i+1).value = column
        
        col = col + 2
    # ファイルの保存
    wb.save(output_file)

txt2xlsx()
